from openpyxl.comments import Comment
from openpyxl.styles import Font
import pandas as pd


def make_excel_report(system_num, excel, result_path, system_copied, inspection_result, detail, solution, stats_column,
                      ip, os, name, all_system):

    system_copied[f'B{3}'] = "▣ " + name + " 상세 진단 결과"

    system_copied[f'F{6}'] = ip
    system_copied[f'F{7}'] = os
    system_copied[f'D{6}'] = name

    num = 0
    # 시스템별 점검 시트_데이터 삽입
    for i in range(39):
        if i in [5, 20, 36]:
            continue

        # 인덱스가 시리즈 범위를 넘지 않는지 확인
        if num < len(inspection_result):
            inspection_cell = f'J{26 + i}'  # "진단결과" 셀의 위치 지정
            solution_cell = f'O{26 + i}'  # "개선방안" 셀의 위치 지정
            system_copied[inspection_cell] = inspection_result[num]
            system_copied[solution_cell] = solution[num]

            detail_cell = f'K{26 + i}'  # "시스템 취약점 현황" 셀의 위치 지정
            if inspection_result[num] == "양호":
                system_copied[detail_cell] = "[시스템 현황]"
            elif inspection_result[num] == "취약":
                system_copied[inspection_cell].font = Font(color='FF0000')
                system_copied[detail_cell] = "[취약점 현황]"
                system_copied[detail_cell].font = Font(color='FF0000')
            elif pd.isna(inspection_result[num]):
                system_copied[inspection_cell] = "N/A"
                system_copied[detail_cell] = "[ 시스템 현황 ]"
            else:  # 인터뷰
                system_copied[inspection_cell] = "인터뷰"
                system_copied[detail_cell] = "[ 시스템 현황 ]"

            system_copied[detail_cell].comment = Comment(detail[num], "Inspector")
        # else:
        #     print(f"Warning: Index {num} is out of bounds for inspection_result with length {len(inspection_result)}")
        
        num += 1

    stats_result = excel['항목별 통계']  # "항목별 통계" 시트
    # 항목별 통계_데이터 삽입
    for i in range(36):
        if i < len(inspection_result):  # 인덱스가 범위를 넘지 않도록 검사
            stats_cell = f'{stats_column}{5 + i}'
            stats_result[stats_cell] = inspection_result[i]
            if pd.isna(inspection_result[i]):
                stats_result[stats_cell] = "N/A"

    sec_level = excel['보안 수준 통계']  # "보안 수준 통계" 시트
    # 시스템 정보 삽입
    insert_row = 9 + (system_num - 1) * 5
    sec_level[f'B{insert_row}'] = name
    sec_level[f'C{insert_row}'] = ip
    sec_level[f'D{insert_row}'] = os

    # system_copied의 F16:L20 영역을 보안 수준 통계 시트에 참조로 삽입
    for i in range(16, 21):
        for j, col in enumerate('FGHIJKL', start=0):  # F는 6번째 열
            cell_value = f'=\'{system_copied.title}\'!{col}{i}'
            sec_level.cell(row=insert_row + (i - 16), column=11 + j).value = cell_value

    inspect_target = excel['점검 대상(자산정보)']  # "점검 대상(자산정보)" 시트
    insert_row = 4 + system_num
    inspect_target[f'B{insert_row}'] = system_num
    inspect_target[f'C{insert_row}'] = name
    inspect_target[f'D{insert_row}'] = ip
    inspect_target[f'E{insert_row}'] = os

    summary_stats = excel['요약 통계']  # "요약 통계" 시트

    for i in range(all_system):  # 16, 17
        cell_value = f'=\'{system_copied.title}\'!L{20}'
        summary_stats.cell(row=27 + system_num, column=17).value = cell_value
        summary_stats.cell(row=27 + system_num, column=16).value = name

    # excel.save(result_path)
